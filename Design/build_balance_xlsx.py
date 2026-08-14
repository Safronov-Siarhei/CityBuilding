"""Builds the balance workbook the game imports from (units/economy/buildings tabs) plus a
read-only summary tab whose formulas show what the numbers actually mean."""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")      # cells the designer edits
DERIVED_FILL = PatternFill("solid", fgColor="EDEDED")    # computed, do not edit
HEADER_FILL = PatternFill("solid", fgColor="2F4F4F")
NOTE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY = Font(name=FONT, size=11)
BLUE = Font(name=FONT, size=11, color="0000FF")          # hand-entered input
BLACK = Font(name=FONT, size=11)
TITLE = Font(name=FONT, bold=True, size=13)
NOTE = Font(name=FONT, size=10, italic=True, color="555555")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------------------------------------------------------------- units ----
units = wb.active
units.title = "units"

unit_cols = [
    ("id", 12, "Ключ строки. Его читает код — не переименовывать."),
    ("display_name", 16, "Как называется в интерфейсе."),
    ("max_health", 12, "Запас здоровья."),
    ("attack_damage", 14, "Урон за удар."),
    ("attack_interval_sec", 18, "Секунд между ударами."),
    ("attack_range_units", 18, "Дальность удара по юниту, м."),
    ("attack_range_structures", 22, "Дальность удара по строению, м."),
    ("walk_speed", 12, "Скорость ходьбы, м/с."),
    ("engage_radius", 14, "Радиус, в котором ищет цель без приказа, м."),
    ("recruit_coins", 14, "Цена найма в монетах (0 — нельзя нанять)."),
    ("upkeep_coins_per_day", 20, "Содержание, монет в день."),
    ("dps", 10, "ВЫЧИСЛЯЕТСЯ: урон в секунду."),
    ("ttk_vs_militia_sec", 20, "ВЫЧИСЛЯЕТСЯ: за сколько секунд убивает ополченца."),
    ("ttk_vs_orc_sec", 18, "ВЫЧИСЛЯЕТСЯ: за сколько секунд убивает орка."),
    ("coins_per_hp", 14, "ВЫЧИСЛЯЕТСЯ: цена найма за единицу здоровья."),
]

for col, (name, width, comment) in enumerate(unit_cols, start=1):
    cell = units.cell(row=1, column=col, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BOX
    units.column_dimensions[get_column_letter(col)].width = width
    units.cell(row=2, column=col).comment = None  # placeholder, comments added below

rows = [
    ["militia", "Ополчение", 12, 5, 1.1, 1.4, 2.6, 1.5, 7, 25, 1],
    ["orc", "Орк", 20, 4, 1.2, 1.4, 3, 1.3, 6, 0, 0],
]

for r, values in enumerate(rows, start=2):
    for c, value in enumerate(values, start=1):
        cell = units.cell(row=r, column=c, value=value)
        cell.font = BLUE
        cell.fill = INPUT_FILL
        cell.border = BOX
        if c in (5, 6, 7, 8):
            cell.number_format = "0.0#"

    # Derived: dps, time-to-kill each way, cost per HP. Militia is row 2, orc row 3.
    dps = units.cell(row=r, column=12, value=f"=IFERROR(D{r}/E{r},0)")
    ttk_militia = units.cell(row=r, column=13, value=f"=IFERROR(CEILING($C$2/D{r},1)*E{r},0)")
    ttk_orc = units.cell(row=r, column=14, value=f"=IFERROR(CEILING($C$3/D{r},1)*E{r},0)")
    coins_per_hp = units.cell(row=r, column=15, value=f"=IFERROR(J{r}/C{r},0)")
    for cell in (dps, ttk_militia, ttk_orc, coins_per_hp):
        cell.font = BLACK
        cell.fill = DERIVED_FILL
        cell.border = BOX
        cell.number_format = "0.00"

units.freeze_panes = "B2"
units.cell(row=5, column=1, value="Жёлтые ячейки — правишь ты. Серые считаются формулами: их можно смотреть, но менять бессмысленно.").font = NOTE
units.cell(row=6, column=1, value="Колонки id и display_name читает код. Порядок колонок не важен — импортёр ищет их по названию, лишние колонки игнорирует.").font = NOTE
units.cell(row=7, column=1, value="Добавить нового юнита = добавить строку. Но код узнает о нём только когда для него появится тип в игре.").font = NOTE

# -------------------------------------------------------------- economy ----
economy = wb.create_sheet("economy")
econ_headers = ["key", "value", "comment"]
widths = [40, 12, 70]
for col, (name, width) in enumerate(zip(econ_headers, widths), start=1):
    cell = economy.cell(row=1, column=col, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BOX
    economy.column_dimensions[get_column_letter(col)].width = width

econ_rows = [
    ("army_max_size", 20, "Общий предел армии по всем группам"),
    ("raid_interval_seconds", 90, "Через сколько секунд портал шлёт следующий набег"),
    ("raid_base_size", 2, "Размер набега на первый день"),
    ("raid_days_per_extra_raider", 3, "Каждые N дней в набеге на одного орка больше"),
    ("raid_max_size", 8, "Потолок размера набега"),
    ("portal_max_health", 320, "Запас прочности портала"),
    ("defence_attack_interval_seconds", 1, "Как часто бьёт здание с защитой (стена/башня/казармы/ворота)"),
    ("defence_attack_range_meters", 6, "Радиус обстрела здания с защитой"),
    ("day_length_seconds", 120, "Длина игрового дня в реальных секундах"),
    ("coins_per_citizen_per_day_at_max_tax", 0.5, "Монет с жителя в день при налоге 100%"),
    ("decay_per_day_at_level1", 0.02, "Ветхость в день у здания 1 уровня; делится на уровень"),
    ("decay_penalty_threshold", 0.7, "С какой ветхости начинает падать производство"),
    ("min_decay_production_multiplier", 0.5, "Во сколько раз падает производство при 100% ветхости"),
    ("repair_cost_fraction", 0.4, "Доля стоимости постройки за полный ремонт"),
    ("wood_per_tree", 5, "Дерева за одно срубленное дерево вручную"),
    ("stone_per_rock", 4, "Камня за один валун вручную"),
    ("building_upgrade_level2_multiplier", 1.6, "Только для формул на вкладке buildings: код читает готовые up2_/up3_"),
    ("building_upgrade_level3_multiplier", 2.8, "Только для формул на вкладке buildings: код читает готовые up2_/up3_"),
]

for r, (key, value, comment) in enumerate(econ_rows, start=2):
    k = economy.cell(row=r, column=1, value=key)
    v = economy.cell(row=r, column=2, value=value)
    c = economy.cell(row=r, column=3, value=comment)
    k.font = BODY
    v.font = BLUE
    v.fill = INPUT_FILL
    c.font = NOTE
    for cell in (k, v, c):
        cell.border = BOX

economy.freeze_panes = "A2"
economy.cell(row=len(econ_rows) + 3, column=1,
             value="Ключи в колонке key читает код: переименование = потерянная настройка (импорт напишет об этом ошибкой).").font = NOTE

# ------------------------------------------------------------ buildings ----
# Numbers only. What a building looks like -- footprint, height, colours, model --
# stays in SetupProject.cs, so nothing here can break geometry.
buildings = wb.create_sheet("buildings")

RESOURCES = ["wood", "stone", "iron", "coal", "gold"]

bld_cols = [
    ("id", 14, None),
    ("display_name", 20, None),
    ("category", 15, None),
    ("cost_wood", 11, None), ("cost_stone", 11, None), ("cost_iron", 10, None),
    ("cost_coal", 10, None), ("cost_gold", 10, None),
    ("citizens_granted", 16, None),
    ("max_workers", 12, None),
    ("produces", 11, None),
    ("production_per_tick", 18, None),
    ("production_interval_sec", 20, None),
    ("max_health", 11, None),
    ("defense", 10, None),
    ("fog_reveal_radius", 16, None),
    ("requires", 13, None),
]
bld_cols += [(f"up2_{res}", 10, "derived") for res in RESOURCES]
bld_cols += [(f"up3_{res}", 10, "derived") for res in RESOURCES]
bld_cols += [
    ("comment", 60, "note"),
    ("total_cost", 12, "derived"),
    ("per_day", 12, "derived"),
    ("payback_days", 14, "derived"),
]

for col, (name, width, _kind) in enumerate(bld_cols, start=1):
    cell = buildings.cell(row=1, column=col, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BOX
    buildings.column_dimensions[get_column_letter(col)].width = width

# id, name, category, cost{res}, citizens, workers, produces, per_tick, interval,
# health, defense, fog, requires, upgrade overrides, comment.
# An upgrade cell is a formula (base cost x the level's multiplier) unless it appears in
# `over`: the Town Hall is free to place so there is nothing to scale, and the iron/coal
# entries are deliberate tech gates rather than a function of the base price.
bld_rows = [
    ("House", "Дом", "Housing", {"wood": 10}, 5, 0, "", 0, 6, 80, 0, 10, "", {}, ""),
    ("Cottage", "Коттедж", "Housing", {"wood": 25, "stone": 8}, 8, 0, "", 0, 6, 120, 0, 10, "House", {},
     "Требует Дом: жильё поприличнее ставится только там, где уже есть базовое."),
    ("TownHall", "Ратуша", "Production", {}, 5, 0, "", 0, 6, 400, 20, 20, "",
     {(2, "wood"): 100, (2, "stone"): 60, (3, "wood"): 220, (3, "stone"): 150, (3, "coal"): 20, (3, "gold"): 40},
     "Ставится бесплатно, поэтому стоимости апгрейдов заданы руками, а не формулой от базовой цены."),
    ("FishermanHut", "Хижина рыбака", "Food", {"wood": 15}, 0, 2, "Food", 2, 6, 90, 0, 12, "", {}, ""),
    ("HunterHut", "Хижина охотника", "Food", {"wood": 15, "stone": 5}, 0, 2, "Food", 2, 6, 90, 0, 14, "", {}, ""),
    ("Farm", "Ферма", "Food", {"wood": 15, "stone": 5}, 0, 3, "Food", 3, 6, 70, 0, 10, "", {}, ""),
    ("Lumberjack", "Лесопилка", "Production", {"wood": 20, "stone": 5}, 0, 3, "Wood", 2, 6, 100, 0, 15, "", {}, ""),
    ("Quarry", "Каменоломня", "Production", {"wood": 20}, 0, 3, "Stone", 2, 6, 110, 0, 14, "", {}, ""),
    ("Mine", "Шахта", "Production", {"wood": 25, "stone": 15}, 0, 3, "Iron", 1, 6, 110, 0, 14, "",
     {(2, "coal"): 10, (3, "coal"): 22},
     "Копать глубже за рудой = топливо для горна и насосов, поэтому в апгрейдах уголь."),
    ("CoalMine", "Угольная шахта", "Production", {"wood": 25, "stone": 10}, 0, 3, "Coal", 2, 6, 100, 0, 14, "",
     {(2, "iron"): 6, (3, "iron"): 14},
     "Железная крепь для глубоких пластов — связывает две шахты друг с другом."),
    ("Wall", "Стена", "Military", {"stone": 5}, 0, 0, "", 0, 6, 150, 15, 6, "",
     {(2, "iron"): 6, (3, "iron"): 14},
     "Железо в апгрейдах у всей линии обороны: сначала цепочка Шахта → железо, потом крепкие стены."),
    ("Tower", "Башня", "Military", {"wood": 5, "stone": 15}, 0, 0, "", 0, 6, 220, 25, 18, "Wall",
     {(2, "iron"): 12, (3, "iron"): 26},
     "Требует Стену: башня усиливает линию, а не стоит сама по себе."),
    ("Barracks", "Казармы", "Military", {"wood": 15, "stone": 30}, 0, 0, "", 0, 6, 180, 10, 10, "",
     {(2, "iron"): 15, (3, "iron"): 32},
     "Оружие и броня для большего гарнизона — самый крупный сток железа в игре."),
    ("Gate", "Ворота", "Military", {"wood": 5, "stone": 12}, 0, 0, "", 0, 6, 160, 12, 8, "",
     {(2, "iron"): 8, (3, "iron"): 18}, ""),
    ("Road", "Дорога", "Infrastructure", {"stone": 3}, 0, 0, "", 0, 6, 40, 0, 4, "", {}, ""),
    ("Bridge", "Мост", "Infrastructure", {"wood": 8}, 0, 0, "", 0, 6, 40, 0, 4, "", {}, ""),
    ("WaterMill", "Водяная мельница", "Food", {"wood": 25, "stone": 10}, 0, 3, "Food", 3, 6, 90, 0, 12, "", {}, ""),
    ("Dock", "Пристань", "Production", {"wood": 20, "stone": 8}, 0, 2, "Gold", 1, 6, 80, 0, 12, "", {}, ""),
]

COST_START = 4                      # column D: cost_wood
UP_START = {2: 18, 3: 23}           # columns R and W
ECON = 'INDEX(economy!B:B,MATCH("{key}",economy!A:A,0))'

def authored(cell, value, fmt=None):
    cell.value = value
    cell.font = BLUE
    cell.fill = INPUT_FILL
    cell.border = BOX
    if fmt:
        cell.number_format = fmt

def derived(cell, formula, fmt="0"):
    cell.value = formula
    cell.font = BLACK
    cell.fill = DERIVED_FILL
    cell.border = BOX
    cell.number_format = fmt

for r, (bid, name, category, cost, citizens, workers, produces, per_tick,
        interval, health, defense, fog, requires, over, comment) in enumerate(bld_rows, start=2):
    authored(buildings.cell(row=r, column=1), bid)
    authored(buildings.cell(row=r, column=2), name)
    authored(buildings.cell(row=r, column=3), category)
    for i, res in enumerate(RESOURCES):
        authored(buildings.cell(row=r, column=COST_START + i), cost.get(res, 0))
    authored(buildings.cell(row=r, column=9), citizens)
    authored(buildings.cell(row=r, column=10), workers)
    authored(buildings.cell(row=r, column=11), produces)
    authored(buildings.cell(row=r, column=12), per_tick)
    authored(buildings.cell(row=r, column=13), interval)
    authored(buildings.cell(row=r, column=14), health)
    authored(buildings.cell(row=r, column=15), defense)
    authored(buildings.cell(row=r, column=16), fog)
    authored(buildings.cell(row=r, column=17), requires)

    for level in (2, 3):
        multiplier = ECON.format(key=f"building_upgrade_level{level}_multiplier")
        for i, res in enumerate(RESOURCES):
            cell = buildings.cell(row=r, column=UP_START[level] + i)
            if (level, res) in over:
                authored(cell, over[(level, res)])
            else:
                base = f"{get_column_letter(COST_START + i)}{r}"
                # Mirrors the rule the code used to apply: scale, round, and never let a
                # non-zero base cost scale down to nothing.
                derived(cell, f"=IF({base}=0,0,MAX(1,ROUND({base}*{multiplier},0)))")

    note_cell = buildings.cell(row=r, column=28, value=comment)
    note_cell.font = NOTE
    note_cell.border = BOX

    derived(buildings.cell(row=r, column=29), f"=SUM(D{r}:H{r})")
    day = ECON.format(key="day_length_seconds")
    derived(buildings.cell(row=r, column=30), f"=IF(M{r}=0,0,J{r}*L{r}*{day}/M{r})", "0.0")
    derived(buildings.cell(row=r, column=31), f'=IF(AD{r}=0,"",AC{r}/AD{r})', "0.0")

buildings.freeze_panes = "B2"
first_note = len(bld_rows) + 3
buildings.cell(row=first_note, column=1,
               value="Жёлтые ячейки — правишь ты. Серые считаются формулами.").font = NOTE
buildings.cell(row=first_note + 1, column=1,
               value="Колонку id читает код: переименование строки = здание без баланса, сборка скажет об этом ошибкой.").font = NOTE
buildings.cell(row=first_note + 2, column=1,
               value="Визуал (размер в клетках, высота, цвета, модель) живёт в SetupProject.cs и в таблицу не переносится.").font = NOTE
buildings.cell(row=first_note + 3, column=1,
               value="per_day и payback_days считают только производство: сколько ресурса в день на полном штате и за сколько дней отбивается постройка.").font = NOTE

# -------------------------------------------------------------- summary ----
s = wb.create_sheet("сводка")
s.column_dimensions["A"].width = 46
s.column_dimensions["B"].width = 16
s.column_dimensions["C"].width = 62

def title(row, text):
    cell = s.cell(row=row, column=1, value=text)
    cell.font = TITLE
    return row + 1

def line(row, label, formula, comment, fmt="0.00"):
    a = s.cell(row=row, column=1, value=label)
    b = s.cell(row=row, column=2, value=formula)
    c = s.cell(row=row, column=3, value=comment)
    a.font = BODY
    b.font = BLACK
    b.fill = DERIVED_FILL
    b.number_format = fmt
    b.border = BOX
    c.font = NOTE
    return row + 1

r = 1
r = title(r, "Дуэль: ополченец против орка")
r = line(r, "Орк убивает ополченца за, сек", "=units!M3", "Меньшее число выигрывает дуэль.")
r = line(r, "Ополченец убивает орка за, сек", "=units!N2", "")
r = line(r, "Кто выигрывает 1 на 1", '=IF(units!M3<units!N2,"орк","ополченец")', "По задумке — орк.", "General")
r = line(r, "Сколько ополченцев нужно, чтобы успеть", "=CEILING(units!N2/units!M3,1)", "Столько бойцов убивают орка быстрее, чем он одного из них.")
r += 1

r = title(r, "Экономика армии")
r = line(r, "Содержание одного бойца, мон./день", "=units!K2", "")
r = line(r, "Содержание полной армии, мон./день", "=units!K2*INDEX(economy!B:B,MATCH(\"army_max_size\",economy!A:A,0))", "При полном кэпе.")
r = line(r, "Налоговый доход при 20 жителях, мон./день", "=20*INDEX(economy!B:B,MATCH(\"coins_per_citizen_per_day_at_max_tax\",economy!A:A,0))", "При налоге 100%.")
r = line(r, "Хватает ли налога на полную армию", '=IF(B10>=B9,"да","нет")', "Если нет — армию придётся держать меньше кэпа или поднимать экономику.", "General")
r = line(r, "Найм полной армии стоит, монет", "=units!J2*INDEX(economy!B:B,MATCH(\"army_max_size\",economy!A:A,0))", "Разовая трата.")
r += 1

r = title(r, "Штурм портала")
r = line(r, "Прочность портала", "=INDEX(economy!B:B,MATCH(\"portal_max_health\",economy!A:A,0))", "")
r = line(r, "Урон группы из 5 ополченцев, в секунду", "=5*units!L2", "")
r = line(r, "5 ополченцев ломают портал за, сек", "=IFERROR(B15/B16,0)", "Без учёта того, что их будут бить.")
r = line(r, "10 ополченцев ломают портал за, сек", "=IFERROR(B15/(10*units!L2),0)", "")
r += 1

r = title(r, "Набеги")
r = line(r, "Набег на 1-й день, орков", "=INDEX(economy!B:B,MATCH(\"raid_base_size\",economy!A:A,0))", "", "0")
r = line(r, "Набег на 10-й день, орков", "=MIN(INDEX(economy!B:B,MATCH(\"raid_max_size\",economy!A:A,0)),INDEX(economy!B:B,MATCH(\"raid_base_size\",economy!A:A,0))+INT(9/INDEX(economy!B:B,MATCH(\"raid_days_per_extra_raider\",economy!A:A,0))))", "", "0")
r = line(r, "Набег на 30-й день, орков", "=MIN(INDEX(economy!B:B,MATCH(\"raid_max_size\",economy!A:A,0)),INDEX(economy!B:B,MATCH(\"raid_base_size\",economy!A:A,0))+INT(29/INDEX(economy!B:B,MATCH(\"raid_days_per_extra_raider\",economy!A:A,0))))", "", "0")
r = line(r, "Набегов за игровой день", "=INDEX(economy!B:B,MATCH(\"day_length_seconds\",economy!A:A,0))/INDEX(economy!B:B,MATCH(\"raid_interval_seconds\",economy!A:A,0))", "Сколько волн приходит за один день календаря.")
r += 1

r = title(r, "Износ зданий")
r = line(r, "Дней до полного износа, 1 уровень", "=1/INDEX(economy!B:B,MATCH(\"decay_per_day_at_level1\",economy!A:A,0))", "После этого здание разрушается.", "0.0")
r = line(r, "Дней до полного износа, 3 уровень", "=3/INDEX(economy!B:B,MATCH(\"decay_per_day_at_level1\",economy!A:A,0))", "Уровень замедляет износ.", "0.0")
r = line(r, "Реальных минут до полного износа, 1 ур.", "=B27*INDEX(economy!B:B,MATCH(\"day_length_seconds\",economy!A:A,0))/60", "", "0.0")

r += 1
r = title(r, "Здания")
r = line(r, "Быстрее всех окупается, дней", "=MIN(buildings!AE:AE)", "По колонке payback_days: сколько игровых дней производство отбивает постройку.", "0.0")
r = line(r, "Линия обороны целиком, камня", '=SUMIF(buildings!$C:$C,"Military",buildings!$E:$E)', "Стена + башня + казармы + ворота по одному разу.", "0")
r = line(r, "Жителей за всё жильё", '=SUMIF(buildings!$C:$C,"Housing",buildings!$I:$I)', "Дом + коттедж по одному разу.", "0")
r = line(r, "Еды в день со всех источников", '=SUMIF(buildings!$C:$C,"Food",buildings!$AD:$AD)', "На полном штате рабочих.", "0.0")

r += 1
note = s.cell(row=r, column=1, value="Эта вкладка только считает — правь units, economy и buildings, числа здесь пересчитаются сами.")
note.font = NOTE

# Next to this script, i.e. Design/ -- the workbook is committed alongside its generator.
wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), "CityBuilding_Balance.xlsx"))
print("saved")
